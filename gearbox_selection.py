# Gearbox Selection
# Start - 04/12/2020

# This program is to take some input information and then make some selections of which gearbox would be appropriate

import openpyxl, logging  # To allow us to use the openpyxl and logging packages and their functions

logging.basicConfig(level=logging.DEBUG, format='%(message)s')  # This will allow you to see all the debugging messages that are debug class or higher
#logging.disable(logging.CRITICAL)  # This will disable all the logging messages that are critical or less

print('\nGearbox Selection Program')

# 1 - Determine what type of gearbox is needed and open the appropriate spreadsheet for it
while True:  # To lop infinitely until the user makes a proper choice for gearbox type
    output_direction = input('\nWould you like the gearbox output to be "right angle", "inline", or "parallel"? ').lower()  # Asks the use what sort of gearbox they are looking for
    # The below statements will open the appropriate spreadsheet with data in it for each type of gearbox
    if output_direction == 'right angle':  # This will occur if you want a right angle gearbox
        gearbox_workbook = openpyxl.load_workbook('Right_Angle_Gearboxes.xlsx')  # Opens the right angle gearbox spreadsheet
        break
    if output_direction == 'inline':
        gearbox_workbook = openpyxl.load_workbook('Inline_Gearboxes.xlsx')  # Opens the inline gearbox spreadsheet
        break
    if output_direction == 'parallel':
        gearbox_workbook = openpyxl.load_workbook('Parallel_Gearboxes.xlsx')  # Opens the parallel gearbox spreadsheet
        break
    else:  # This will occur if none of the types selected exist, or if the user inputs something other than the 3 options available
        print('You will need to type in "right angle", "inline", or "parallel"')
        continue

logging.debug(f'\nWorkbook opened: {gearbox_workbook}')

# 2 - Find out the parameters that need to be met for the gearbox.
# 2.1 - Finding the input motor speed and poles
while True:  # To infinitely loop through this until an acceptable answer is given
    try:
        motor_input_speed = int(input('What is the motor speed (rpm)? Enter 0 to input motor poles. '))  # To take an input of a motor speed
        # The below if statements will look for the motor speed and determine how many poles the motor is
        if 2700 <= motor_input_speed <= 3000:
            motor_poles_required = 2
            break
        elif 1350 <= motor_input_speed <= 1500:
            motor_poles_required = 4
            break
        elif 900 <= motor_input_speed <= 1000:
            motor_poles_required = 6
            break
        elif motor_input_speed == 0:  # To try for an input of motor poles instead.
            motor_poles_required = int(input('How many poles is your motor? 2, 4, or 6? or 0 for all options '))
            # The below if statements will assign an arbitrary rough motor speed based on the number of poles input
            if motor_poles_required == 2:
                motor_input_speed = 2800
                break
            elif motor_poles_required == 4:
                motor_input_speed = 1400
                break
            elif motor_poles_required == 6:
                motor_input_speed = 900
                break
            elif motor_poles_required == 0:
                break
            else:
                print('Motor poles needs to be 2, 4, 6, or 0. Please try again.\n')
                continue
        else:
            print('Motor input speed needs to be in a reasonable range. Please try again.\n')
            continue
    except ValueError:
        print('Motor speed or pole numbers will need to be an integer. Please try again.\n')
        continue

logging.debug(f'Motor speed: {motor_input_speed} rpm')
logging.debug(f'Motor poles: {motor_poles_required}')

# 2.2 - Finding the output speed or ratio required
while True:  # To infinitely loop until a valid answer is given
    try:  # Will attempt the below unless an exception is thrown
        output_speed_required = int(input('What is the output speed required (rpm)? Enter 0 to input a ratio. '))  # Asks the user about the output speed or ratio
        #  The below statements will look for an output speed or a ratio
        if output_speed_required > 0:  # This will occur if the output speed is 0 or less than 0
            ratio_required = motor_input_speed / output_speed_required
            break
        elif output_speed_required == 0:
            ratio_required = int(input('What ratio would you like? '))
            # These statements will look for a ratio
            if ratio_required > 0:
                output_speed_required = motor_input_speed / ratio_required
                break
            else:
                print('The ratio needs to be positive and above 0. Please try again.\n')
                continue
        else:
            print('You\'ve put in a negative number. Please find a proper value\n')
            continue
    except ValueError:  # Will occur if a non integer is input
        print('This needs to be an integer, try again\n')
        continue

logging.debug(f'\nOutput Speed: {round(output_speed_required, 1)} rpm')
logging.debug(f'Ratio: {round(ratio_required, 1)} : 1')

# 2.2 - Finding the input power or output torque required
while True:  # To infinitely loop until a valid answer is given
    try:  # Will attempt the below unless an exception is thrown
        input_power = float(input('What is the motor power input (kW)? enter 0 to input torque instead. '))  # Asks the user about the power required
        # The below statements will look for an input power or leave it
        if input_power > 0:  # This will occur if the power entered is not 0 or less than 0
            output_torque_required = 9550 * input_power / output_speed_required  # This will change the power input into a torque number, assuming no losses
            break
        elif input_power < 0:
            print('You\'ve put it in as a negative number. Please find a proper value\n')
            continue
        elif input_power == 0:
            output_torque_required = int(input('What is the torque required (Nm)?'))  # Asks the user about the torque required
            # These statements will look for an output torque
            if output_torque_required > 0:  # This will occur if the torque entered is not 0 or less than 0
                input_power = output_torque_required * output_speed_required / 9550
                break
            elif output_torque_required < 0:
                print('You\'ve put it in as a negative number. Please find a proper value\n')
                continue
            elif output_torque_required == 0:
                print('You need to put in a power or a torque. Please try again\n')
                continue
    except ValueError:  # Will occur if a non integer is input
        print('This needs to be an integer, try again\n')
        continue

logging.debug(f'\nOutput Torque Required: {round(output_torque_required, 1)} Nm')
logging.debug(f'Input power: {round(input_power, 2)} kW')

# 2.3 - Finding the shaft size required
while True:
    try:
        shaft_size_min = int(input('What is the minimum shaft size you can work with? (mm) '))
        shaft_size_max = int(input('What is the maximum shaft size you can work with? (mm) '))
        if shaft_size_min > shaft_size_max:
            print(f'Minimum value needs to be below maximum value. Try again\n')
            continue
        elif shaft_size_min < 0 or shaft_size_max < 0:
            print(f'These both need to be above 0. Try again\n')
            continue
        else:
            break
    except ValueError:
        print('These values need to be an integers. Try again\n')
        continue

logging.debug(f'\nShaft size range: {shaft_size_min} mm   to   {shaft_size_max} mm')

# 2.4 - Finding the tolerances and range we can work with for various inputs
# This is the output speed and ratio tolerance
while True:
    try:
        output_speed_tolerance = int(input('What output speed or Ratio tolerance do you want to work with? (%) '))
        if output_speed_tolerance > 0:
            output_speed_tolerance = output_speed_tolerance / 100  # This will turn this into a decimal
            break
        else:
            print('This needs to be positive and above 0.  Try again.\n')
            continue
    except ValueError:
        print('This needs to be an integer. Try again\n')
        continue

logging.debug(f'Speed/Ratio Tolerance = {output_speed_tolerance * 100} %')

# This is the output torque or input power tolerance
while True:
    try:
        output_torque_tolerance = int(input('What output torque tolerance or input torque do you want to work with? (%) '))
        if output_torque_tolerance > 0:
            output_torque_tolerance = output_torque_tolerance / 100  # This will turn this into a decimal
            break
        else:
            print('This needs to be positive and above 0.  Try again.\n')
            continue
    except ValueError:
        print('This needs to be an integer. Try again\n')
        continue

logging.debug(f'Output torque Tolerance = {output_torque_tolerance * 100} %')

# 3 - Making an initial list of all the applicable series
series_options_list = gearbox_workbook.sheetnames  # To create a base list of appropriate options

logging.debug(f'\nAll gearbox options: \n{series_options_list}')
logging.debug(f'List length: {len(series_options_list)}')

for gearbox_sheet_name in series_options_list:  # This will iterate through all the sheets in the workbook
    gearbox_sheet = gearbox_workbook[gearbox_sheet_name]  # This will set the sheet we are looking in to the current sheet
    # These below lines will find some information from the worksheets given all the data from the user
    series_max_rated_torque = gearbox_sheet['B1'].value  # This is the max rated torque for this size gearbox
    series_shaft_size_1 = gearbox_sheet['L1'].value  # This is the normal shaft/bore size for this gearbox

    # These lines will exclude all the options from the list that do not have all their base data filled in
    if gearbox_sheet_name is None or series_max_rated_torque is None or series_shaft_size_1 is None:
        try:
            series_options_list.remove(gearbox_sheet_name)  # This wil remove unacceptable options from the list
        except ValueError:  # This will occur if the series has already been removed from the list
            pass

logging.debug(f'\nOptions with base data: \n{series_options_list}')
logging.debug(f'List length: {len(series_options_list)}')

# 4 - Narrowing the applicable series options
compatible_series_list = []
for gearbox_sheet_name in series_options_list:
    gearbox_sheet = gearbox_workbook[gearbox_sheet_name]  # This will set the sheet we are looking in to the current sheet
    # These below lines will find some information from the worksheets given all the data from the user
    series_max_rated_torque = gearbox_sheet['B1'].value  # This is the max rated torque for this size gearbox
    series_shaft_size_1 = gearbox_sheet['L1'].value  # This is the normal shaft/bore size for this gearbox
    series_shaft_size_2 = gearbox_sheet['M1'].value  # THis will be an alternate value for the shaft/bore size if there is one

    # This will filter all the good series options down into their own list
    if output_torque_required <= series_max_rated_torque:
        if series_shaft_size_2 is not None:
            if shaft_size_min <= series_shaft_size_1 and shaft_size_min <= series_shaft_size_2:
                if shaft_size_max >= series_shaft_size_1 and shaft_size_max >= series_shaft_size_2:
                    compatible_series_list.append(gearbox_sheet_name)

        elif series_shaft_size_2 is None:
            if shaft_size_min < series_shaft_size_1:
                if shaft_size_max > series_shaft_size_1:
                    compatible_series_list.append(gearbox_sheet_name)

logging.debug(f'\nGood options: \n{compatible_series_list}')
logging.debug(f'List length: {len(compatible_series_list)}')

# 5 - Narrowing down the motor pole options to look through
if motor_poles_required == 2:
    speed_columns = ['K']
elif motor_poles_required == 4:
    speed_columns = ['H']
elif motor_poles_required == 6:
    speed_columns = ['E']
elif motor_poles_required == 0:
    speed_columns = ['K', 'H', 'E']

logging.debug(f'\nMotor Poles: {motor_poles_required}')
logging.debug(f'speed columns: {speed_columns}')

acceptable_gearbox_list = [[], [], [], [], [], []]

# 6 - Finding all the sheet data for all the applicable options so far
for compatible_series in compatible_series_list:
    gearbox_sheet = gearbox_workbook[compatible_series]
    for sheet_column in speed_columns:
        for sheet_row in range(4, gearbox_sheet.max_row + 1):
            sheet_output_speed = gearbox_sheet[sheet_column + str(sheet_row)].value  # This will go through all the speed values in the sheet
            sheet_output_speed_coordinates = [gearbox_sheet[sheet_column + str(sheet_row)].row, gearbox_sheet[sheet_column + str(sheet_row)].column]
            sheet_max_output_torque = gearbox_sheet.cell(row=sheet_output_speed_coordinates[0], column=sheet_output_speed_coordinates[1] + 1).value
            sheet_max_input_power = gearbox_sheet.cell(row=sheet_output_speed_coordinates[0], column=sheet_output_speed_coordinates[1] + 2).value
            sheet_gearbox_ratio = gearbox_sheet.cell(row=sheet_output_speed_coordinates[0], column=1).value
            sheet_motor_poles = gearbox_sheet.cell(row=2, column=sheet_output_speed_coordinates[1]).value

# 6.1 - Only including the options within the acceptable range
            if sheet_output_speed is not None and sheet_max_output_torque is not None and sheet_max_input_power is not None:  # This will exclude the sheets that have not been filled in
                if output_speed_required * (1 - output_speed_tolerance) <= sheet_output_speed <= output_speed_required * (1 + output_speed_tolerance):
                    if output_torque_required * (1 - output_torque_tolerance) <= sheet_max_output_torque:  # <= output_torque_required * (1 + output_torque_tolerance):
                        try:
                            acceptable_gearbox_list[0].append(sheet_output_speed)
                            acceptable_gearbox_list[1].append(sheet_max_output_torque)
                            acceptable_gearbox_list[2].append(sheet_max_input_power)
                            acceptable_gearbox_list[3].append(sheet_gearbox_ratio)
                            acceptable_gearbox_list[4].append(sheet_motor_poles)
                            acceptable_gearbox_list[5].append(compatible_series)
                        except TypeError:
                            pass

logging.debug(f'\n{acceptable_gearbox_list}')

# 7 - Displaying all the options that have been found
if len(acceptable_gearbox_list[0]) >= 1:
    print('\nGearbox selections are:\n')
    for gearbox_option in range(len(acceptable_gearbox_list[0])):
        gearbox_output_speed = acceptable_gearbox_list[0][gearbox_option]
        gearbox_max_output_torque = acceptable_gearbox_list[1][gearbox_option]
        gearbox_max_input_power = acceptable_gearbox_list[2][gearbox_option]
        gearbox_ratio = acceptable_gearbox_list[3][gearbox_option]
        gearbox_motor_poles = acceptable_gearbox_list[4][gearbox_option]
        gearbox_series = acceptable_gearbox_list[5][gearbox_option]

        actual_output_speed = motor_input_speed / gearbox_ratio
        actual_output_torque = 9550 * input_power / actual_output_speed
        safety_factor = gearbox_max_output_torque / actual_output_torque

        print(f'{gearbox_series}_ {gearbox_ratio}:1 with a {gearbox_motor_poles} motor.')
        print(f'Catalogue Specs for this would be: {gearbox_output_speed}rpm output speed, {gearbox_max_output_torque}Nm max output torque, {gearbox_max_input_power}kW max input power')
        print(f'Actual Specs for this would be: {round(actual_output_speed, 1)}rpm output, {round(actual_output_torque, 1)}Nm output torque, {round(safety_factor, 2)} Safety Factor\n')

else:
    print('\nThere were no appropriate gearbox selections found for the given data.')

# I think this is actually enough for a first attempt.
# End - 21/12/2020

# FUTURE WORK OR FURTHER ITERATIONS
# Interpolate the tables to provide more accurate output speeds and torques
# Provide more useful and accurate information on safety factors
# Make the motor powers you can choose from only actual standard motor powers

# Put everything in functions and classes instead of this sequential spaghetti mess
# Make this work with a GUI, no one else is gonna use a console application
